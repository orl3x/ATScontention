import openpyxl as xl
from openpyxl.styles import Font, Alignment
import datetime
import os


path = "C:/Users/Pruebas/Desktop/REPORTES/"

name = datetime.datetime.today().strftime("%d-%m-%Y")+".xlsx"
date = datetime.datetime.today().strftime("%d-%m-%Y")

def createNewWorkbook():
    os.system("taskkill /im EXCEL.EXE -f")
    global name
    wb = xl.Workbook()
    ws = wb.active
    ws.title = "Data Base"

    #DATA BASE SHEET CELL CONFIGURATION
    reportWs = wb.create_sheet("Resume")
    ws.column_dimensions["A"].width=15
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 15

    #REPORT SHEET CELL CONFIGURATION
    reportWs.column_dimensions["A"].width = 25
    reportWs.merge_cells("A1:B1")
    reportWs["A1"] = "Daily Resume"
    reportWs["A1"].alignment = Alignment(horizontal="center")
    reportWs["A1"].font = Font(bold=True, size=14)

    #MAIN TITLE
    ws.merge_cells("A1:F1")
    ws["A1"] = "DAILY REPORT - ATS1 CHECKER             DATE: "+date
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="center")

    #TIME COLUMN
    ws["A4"] = "TIME"
    ws["A4"].font = Font(bold=True)

    #MODEL COLUMN
    ws["B4"] = "MODEL"
    ws["B4"].font = Font(bold=True)

    #SERIAL COLUMN
    ws["C4"] = "SERIAL"
    ws["C4"].font = Font(bold=True)

    #STATUS COLUMN
    ws["D4"] = "STATUS"
    ws["D4"].font = Font(bold=True)

    #INFO COLUMN
    ws["E4"] = "INFO"
    ws["E4"].font = Font(bold=True)

    # 2ND TIME COLUMN
    ws["F4"] = "2ND TIME"
    ws["F4"].font = Font(bold=True)

    # PASSED 1ST ATTEMPT
    reportWs["A3"] = "PASSED AT 1° ATTEMPT:"
    reportWs["A3"].font = Font(bold=True)

    #PASSED 1ST ATTEMPT RESULT
    reportWs["B3"] = '=COUNTIF(\'Data Base\'!D:D,"PASSED")'
    reportWs["B3"].font = Font(bold=True, color='259e15')

    # FAILED 1ST ATTEMPT
    reportWs["A4"] = "FAILED AT 1° ATTEMPT:"
    reportWs["A4"].font = Font(bold=True)

    #FAILED 1ST ATTEMPT RESULT
    reportWs["B4"] = '=COUNTIF(\'Data Base\'!D:D,"FAILED")'
    reportWs["B4"].font = Font(bold=True, color='cf1729')

    #PASSED 2ND ATTEMPT
    reportWs["A6"] = "PASSED IN 2° ATTEMPT:"
    reportWs["A6"].font = Font(bold=True)

    #PASSED 2ND ATTEMPT RESULT
    reportWs["B6"] = '=COUNTIF(\'Data Base\'!F:F,"PASSED")'
    reportWs["B6"].font = Font(bold=True, color='259e15')

    # TOTAL PASSED
    reportWs["A8"] = "TOTAL PASSED UNITS:"
    reportWs["A8"].font = Font(bold=True)

    # TOTAL PASSED RESULT
    reportWs["B8"] = '=B3+B6'
    reportWs["B8"].font = Font(bold=True, color='259e15')

    #TOTAL FAILED
    reportWs["A9"] = "TOTAL FAILED UNITS:"
    reportWs["A9"].font = Font(bold=True)

    # TOTAL FAILED RESULT
    reportWs["B9"] = '=B4-B6'
    reportWs["B9"].font = Font(bold=True, color='cf1729')

    #TOTAL UNITS PROCESSED
    reportWs["A11"] = "TOTAL UNITS PROCESSED:"
    reportWs["A11"].font = Font(bold=True,color='1410fe')

    # TOTAL UNITS PROCESSED RESULT
    reportWs["B11"] = '=B8+B9'
    reportWs["B11"].font = Font(bold=True)



    #SAVE WORKBOOK
    saveWB(wb)


def saveWB(wb):
    wb.save("C:/Users/Pruebas/Desktop/REPORTES/"+name)
    print("Saved succesfully")

def isFileExisting():
    global name
    global path
    return os.path.isfile(path+name)

def writeToWorkbook(model, serial, status):
    os.system("taskkill /im EXCEL.EXE -f")

    ## DEFINE STATUS LABEL
    if status == 1:
        statusLabel = "PASSED"
        font = "Font(color='259e15')"

    else:
        statusLabel = "FAILED"
        font = "Font(color='cf1729')"

    if status == 1:
        infoLabel = " "
    elif status == 2:
        infoLabel = "FAILED IN ATS1"
    else:
        infoLabel = "NO ATS1"
    ##############################################
    #CHECK IF REPORT EXISTS OR CREATE NEW ONE
    global name
    if isFileExisting():
        wb = xl.load_workbook(path+name)
    else:
        createNewWorkbook()

    wb = xl.load_workbook(path+name)
    ws = wb.active

    #SEARCH FOR EXISTING SERIALS
    searchResult = isSerialFound(serial, ws)

    #IF SERIAL NOT FOUND, WRITE IT TO BOOK
    if searchResult == -1:
        currentTime = datetime.datetime.now().strftime("%H:%M:%Shrs")
        print("Serial Guardado")
        print(currentTime)
        timeRow = "A" + str(ws.max_row + 1)
        print(timeRow)
        modelRow = "B" + str(ws.max_row + 1)
        print(modelRow)
        serialRow = "C" + str(ws.max_row + 1)
        print(serialRow)
        statusRow = "D" + str(ws.max_row + 1)
        print(statusRow)
        infoRow = "E" + str(ws.max_row + 1)
        print(infoRow)
        ws[timeRow] = str(currentTime)
        ws[modelRow] = str(model)
        ws[serialRow] = str(serial)

        ws[statusRow] = statusLabel
        ws[statusRow].font = eval(font)

        ws[infoRow] = infoLabel


        saveWB(wb)
        print("Files saved succesfully")

    elif searchResult > -1:
        foundRow = "D"+str(searchResult)
        print("Found row is "+foundRow)
        if ws[foundRow].value == "FAILED" and status == 1:
            print("inside if")
            secondTimeCell = "F" + str(searchResult)
            ws[secondTimeCell] = "PASSED"
            ws[secondTimeCell].font = Font(color='259e15')
            print("Second test passed")
            saveWB(wb)
        else:
            print("Serial Repeated")


def isSerialFound(serial, ws):
    for row in range(3,ws.max_row+1):
        currentRow = "C"+str(row)
        if ws[currentRow].value == str(serial):
            print("Found")
            return row
    print("Not Found")
    return -1

